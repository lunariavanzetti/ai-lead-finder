from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.exports.common import lead_to_flat_dict
from app.models.lead import Lead


def export_xlsx(leads: list[Lead], out_path: Path) -> Path:
    rows = [lead_to_flat_dict(lead) for lead in leads]
    df = pd.DataFrame(rows)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        sheet = writer.sheets["Leads"]

        for col_idx, column in enumerate(df.columns, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            max_len = max([len(str(column))] + [len(str(v)) for v in df[column].astype(str)])
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    return out_path
